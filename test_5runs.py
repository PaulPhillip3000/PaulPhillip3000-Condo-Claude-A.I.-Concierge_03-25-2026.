#!/usr/bin/env python3
"""
Run ground-truth generation 5 times and verify:
  1. Demand letter table has NO blank cells — every row shows $X.XX or $0.00
  2. Maintenance/assessments is non-zero
  3. Excel Sheet 1 (Statement of Account) contains the SAME 9-row table
     with dollar-formatted values matching the demand letter
  4. Consistent across all 5 runs
"""
import asyncio
import sys
import os
import re
from pathlib import Path

sys.path.insert(0, "/workspace")
os.chdir("/workspace")
Path("/workspace/outputs").mkdir(exist_ok=True)


def read_soa_from_excel(xl_path: str) -> list:
    """Read Sheet 1 SOA rows as (label, value_string) pairs."""
    import openpyxl
    wb = openpyxl.load_workbook(xl_path, data_only=True)
    if "Statement of Account" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Statement of Account"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        label = str(row[0] or "").strip()
        val = str(row[1] or "").strip()
        if not label:
            continue
        # Skip non-table rows (notes, verification, etc.)
        if any(kw in label.lower() for kw in ("nola date:", "cross-sheet", "soa total",
                                                "nola-ledger", "variance", "status")):
            continue
        rows.append((label, val))
    wb.close()
    return rows


def read_demand_letter_table(docx_path: str) -> list:
    """Read the demand letter table rows as (label, value_string) pairs."""
    from docx import Document
    doc = Document(docx_path)
    rows = []
    for table in doc.tables:
        for row in table.rows:
            if len(row.cells) < 2:
                continue
            label = row.cells[0].text.strip()
            val = row.cells[1].text.strip()
            rows.append((label, val))
    return rows


def parse_dollar(s: str) -> float:
    """Parse dollar string to float."""
    s = s.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    try:
        return round(float(s), 2)
    except (ValueError, TypeError):
        return None


async def run_once():
    from backend import generate_ground_truth
    return await generate_ground_truth(certified_mail=40.0, other_costs=16.0, attorney_fees=400.0)


async def main():
    RUNS = 5
    all_results = []
    failures = []

    print("=" * 70)
    print(f"  RUNNING {RUNS} GROUND-TRUTH GENERATIONS")
    print("=" * 70)

    for i in range(1, RUNS + 1):
        print(f"\n{'─' * 60}")
        print(f"  RUN {i}/{RUNS}")
        print(f"{'─' * 60}")

        # Clear previous outputs
        for f in Path("/workspace/outputs").glob("*"):
            f.unlink()

        try:
            result = await run_once()
        except Exception as e:
            msg = f"Run {i}: generation FAILED: {e}"
            print(f"  FAIL: {msg}")
            failures.append(msg)
            import traceback; traceback.print_exc()
            continue

        if result.get("status") != "success":
            msg = f"Run {i}: status={result.get('status')}"
            print(f"  FAIL: {msg}")
            failures.append(msg)
            continue

        xl_files = sorted(Path("/workspace/outputs").glob("*.xlsx"),
                          key=lambda p: p.stat().st_mtime, reverse=True)
        docx_files = sorted(Path("/workspace/outputs").glob("*FirstDemand*.docx"),
                            key=lambda p: p.stat().st_mtime, reverse=True)

        if not xl_files or not docx_files:
            msg = f"Run {i}: Missing output files (xlsx={len(xl_files)}, docx={len(docx_files)})"
            print(f"  FAIL: {msg}")
            failures.append(msg)
            continue

        # ── Read both tables ──
        soa_rows = read_soa_from_excel(str(xl_files[0]))
        dl_rows = read_demand_letter_table(str(docx_files[0]))

        print(f"\n  EXCEL SHEET 1 — Statement of Account:")
        for label, val in soa_rows:
            blank = " *** BLANK ***" if not val or val in ("[BLANK]", "None") else ""
            print(f"    {label:55s}  {val}{blank}")

        print(f"\n  FIRST DEMAND LETTER — Amount Due Summary:")
        for label, val in dl_rows:
            blank = " *** BLANK ***" if not val or val in ("[BLANK]", "None") else ""
            print(f"    {label:55s}  {val}{blank}")

        # ── CHECK 1: No blank cells in demand letter table ──
        dl_blanks = [(l, v) for l, v in dl_rows if not v or v.strip() == ""]
        if dl_blanks:
            msg = f"Run {i}: Demand letter has BLANK cells: {[l for l,v in dl_blanks]}"
            print(f"\n  FAIL: {msg}")
            failures.append(msg)
        else:
            print(f"\n  OK: No blank cells in demand letter table ({len(dl_rows)} rows)")

        # ── CHECK 2: No blank cells in SOA ──
        soa_blanks = [(l, v) for l, v in soa_rows if not v or v.strip() == ""]
        if soa_blanks:
            msg = f"Run {i}: SOA has BLANK cells: {[l for l,v in soa_blanks]}"
            print(f"\n  FAIL: {msg}")
            failures.append(msg)
        else:
            print(f"  OK: No blank cells in SOA ({len(soa_rows)} rows)")

        # ── CHECK 3: Maintenance is non-zero ──
        dl_maint_val = None
        soa_maint_val = None
        for l, v in dl_rows:
            if "maintenance" in l.lower():
                dl_maint_val = parse_dollar(v)
        for l, v in soa_rows:
            if "maintenance" in l.lower():
                soa_maint_val = parse_dollar(v)

        if dl_maint_val is None or dl_maint_val == 0:
            msg = f"Run {i}: Demand letter maintenance = ${dl_maint_val or 0:.2f} (should be non-zero)"
            print(f"  FAIL: {msg}")
            failures.append(msg)
        else:
            print(f"  OK: Demand letter maintenance = ${dl_maint_val:,.2f}")

        if soa_maint_val is None or soa_maint_val == 0:
            msg = f"Run {i}: SOA maintenance = ${soa_maint_val or 0:.2f} (should be non-zero)"
            print(f"  FAIL: {msg}")
            failures.append(msg)
        else:
            print(f"  OK: SOA maintenance = ${soa_maint_val:,.2f}")

        # ── CHECK 4: SOA table matches demand letter table (same rows, same values) ──
        if len(soa_rows) != len(dl_rows):
            msg = f"Run {i}: Row count mismatch — SOA={len(soa_rows)} vs DL={len(dl_rows)}"
            print(f"  FAIL: {msg}")
            failures.append(msg)
        else:
            mismatches = []
            for (sl, sv), (dl, dv) in zip(soa_rows, dl_rows):
                # Compare dollar values (parse both)
                sv_f = parse_dollar(sv)
                dv_f = parse_dollar(dv)
                if sv_f is not None and dv_f is not None:
                    if abs(sv_f - dv_f) > 0.01:
                        mismatches.append(f"{sl}: SOA={sv} vs DL={dv}")
                elif sv != dv:
                    mismatches.append(f"{sl}: SOA='{sv}' vs DL='{dv}'")

            if mismatches:
                msg = f"Run {i}: SOA/DL value mismatches: {'; '.join(mismatches)}"
                print(f"  FAIL: {msg}")
                failures.append(msg)
            else:
                print(f"  OK: SOA and Demand Letter tables MATCH (all {len(soa_rows)} rows)")

        # ── CHECK 5: All values are dollar-formatted (no raw numbers) ──
        soa_no_dollar = [(l, v) for l, v in soa_rows if v and "$" not in v and v != "$0.00"]
        if soa_no_dollar:
            msg = f"Run {i}: SOA has non-dollar-formatted values: {soa_no_dollar[:3]}"
            print(f"  FAIL: {msg}")
            failures.append(msg)
        else:
            print(f"  OK: All SOA values are dollar-formatted")

        all_results.append({"soa": soa_rows, "dl": dl_rows})

    # ── Cross-run consistency ──
    print(f"\n{'=' * 70}")
    print("  CROSS-RUN CONSISTENCY")
    print(f"{'=' * 70}")
    if len(all_results) >= 2:
        ref_soa = all_results[0]["soa"]
        for idx, r in enumerate(all_results[1:], 2):
            for (rl, rv), (cl, cv) in zip(ref_soa, r["soa"]):
                rv_f = parse_dollar(rv)
                cv_f = parse_dollar(cv)
                if rv_f is not None and cv_f is not None and abs(rv_f - cv_f) > 0.01:
                    msg = f"Cross-run: Run 1 vs Run {idx} differ on '{rl}': {rv} vs {cv}"
                    print(f"  WARN: {msg}")
                    failures.append(msg)
        if not any("Cross-run" in f for f in failures):
            print("  OK: All runs produced consistent values")
    else:
        print("  SKIP: Not enough successful runs")

    # ── SUMMARY ──
    print(f"\n{'=' * 70}")
    if failures:
        print(f"  RESULT: {len(failures)} FAILURE(S)")
        for f in failures:
            print(f"    - {f}")
    else:
        print(f"  RESULT: ALL {RUNS} RUNS PASSED")
    print(f"{'=' * 70}")
    return len(failures) == 0


if __name__ == "__main__":
    success = asyncio.run(main())
    sys.exit(0 if success else 1)
