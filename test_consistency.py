#!/usr/bin/env python3
"""
CondoClaw Consistency Test — runs generate_ledger + generate_first_letter
for each uploaded matter, then checks Sheet 1 / Sheet 2 / Demand Letter totals.
"""
import asyncio, os, sys, re, traceback, glob, shutil
from pathlib import Path

sys.path.insert(0, "/workspace")
os.chdir("/workspace")
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from docx import Document


def _sf(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    try: return float(str(val).replace("$","").replace(",","").strip())
    except: return None


def extract_soa(wb):
    ws = wb["Statement of Account"]
    total = None; items = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        d, v = str(row[0] or ""), _sf(row[1])
        if v is None: continue
        dl = d.lower()
        if "total" in dl and ("outstanding" in dl or "amount" in dl): total = v
        elif "nola opening" in dl: items["nola_opening"] = v
        elif "assessments" in dl and "post-nola" in dl: items["post_nola_asmts"] = v
        elif "maintenance" in dl and "due" in dl: items["maintenance"] = v
        elif "interest" in dl and ("accrued" in dl or "18%" in dl): items["interest"] = v
        elif "late fee" in dl: items["late_fees"] = v
        elif "attorney" in dl and "other" not in dl and "fee" in dl: items["atty_fees"] = v
        elif "other charges" in dl: items["other_charges"] = v
        elif "credit" in dl or "payment" in dl or "partial" in dl: items["credits"] = v
        elif "certified" in dl: items["certified_mail"] = v
        elif "other cost" in dl or "other attorney" in dl: items["other_costs"] = v
    return total, items


def extract_nl(wb):
    """Extract totals from NOLA-Ledger or Account Ledger (ground-truth endpoint).

    Handles both sheet structures:
    - generate_ledger: "NOLA-Ledger" with NOLA anchor, CURRENT BALANCE, TOTALS
    - generate_ground_truth: "Account Ledger" with AR rows + TOTALS row
    """
    # Try both sheet names
    ws = None
    for name in ("NOLA-Ledger", "Account Ledger"):
        if name in wb.sheetnames:
            ws = wb[name]; break
    if not ws:
        return 0.0, {}, None, None

    is_gt = ws.title == "Account Ledger"

    if is_gt:
        # Ground-truth format: find the TOTALS row and read balance from column L (12)
        # Columns: #, Date, Ref, Desc, Type, Assessments(F), LateFees(G), Interest(H),
        #          AttyFees(I), Other(J), Credits(K), Balance(L), DaysPastDue(M)
        totals_balance = 0.0
        cats = {"assessments":0,"interest":0,"late_fees":0,"atty_fees":0,
                "other":0,"credits":0,"payments":0}
        for row in ws.iter_rows(min_row=2, values_only=False):
            desc = str(row[3].value or "") if len(row) > 3 else ""
            if desc.strip() == "TOTALS":
                # F=6(assess), G=7(late), H=8(int), I=9(atty), J=10(other), K=11(credits), L=12(balance)
                cats["assessments"] = float(row[5].value or 0) if isinstance(row[5].value, (int,float)) else 0
                cats["late_fees"] = float(row[6].value or 0) if isinstance(row[6].value, (int,float)) else 0
                cats["interest"] = float(row[7].value or 0) if isinstance(row[7].value, (int,float)) else 0
                cats["atty_fees"] = float(row[8].value or 0) if isinstance(row[8].value, (int,float)) else 0
                cats["other"] = float(row[9].value or 0) if isinstance(row[9].value, (int,float)) else 0
                cats["credits"] = float(row[10].value or 0) if isinstance(row[10].value, (int,float)) else 0
                totals_balance = float(row[11].value or 0) if isinstance(row[11].value, (int,float)) else 0
                break
        return totals_balance, cats, None, None

    # Original NOLA-Ledger format
    nola_row = None; cb_row = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        d = str(row[2].value or ""); t = str(row[3].value or "")
        if "Balance per Notice of Late Assessment" in d: nola_row = row[0].row
        elif t == "Current Balance": cb_row = row[0].row

    if not nola_row:
        return 0.0, {}, None, None

    running = 0.0
    cats = {"assessments":0,"interest":0,"late_fees":0,"atty_fees":0,
            "other":0,"payments":0,"credits":0}

    # New 15-col: E=Maint(4), F=Special(5), G=Water(6), H=Int(7), I=Late(8),
    #             J=Legal(9), K=Other(10), L=Pay(11), M=Cred(12), N=RunBal(13)
    pre_nola_sums = [0.0]*9  # 7 charge + 2 credit columns
    for row in ws.iter_rows(min_row=2, max_row=nola_row-1, values_only=False):
        if str(row[3].value or "") == "Separator": continue
        for i in range(9):
            v = row[4+i].value
            if isinstance(v, (int,float)): pre_nola_sums[i] += float(v)

    _stop_row = cb_row if cb_row else 9999
    for row in ws.iter_rows(min_row=nola_row, values_only=False):
        rn = row[0].row; d = str(row[2].value or ""); t = str(row[3].value or "")
        if rn >= _stop_row: break
        if t in ("Separator",): continue
        if d.strip() == "TOTALS": break

        vals = []
        for i in range(9):  # E(4) through M(12)
            cell = row[4+i]; v = cell.value
            if isinstance(v, (int,float)): vals.append(float(v))
            elif rn == nola_row and i < len(pre_nola_sums): vals.append(pre_nola_sums[i])
            else: vals.append(0.0)

        charge = sum(vals[:7]); credit = sum(vals[7:9])
        running = round(running + charge - credit, 2)
        cats["assessments"] += vals[0]+vals[1]+vals[2]  # Maint+Special+Water
        cats["interest"] += vals[3]; cats["late_fees"] += vals[4]
        cats["atty_fees"] += vals[5]; cats["other"] += vals[6]
        cats["payments"] += vals[7]; cats["credits"] += vals[8]

    return running, cats, nola_row, cb_row


def extract_dl(docx_path):
    doc = Document(str(docx_path))
    data = {}; total = None
    for table in doc.tables:
        for row in table.rows:
            if len(row.cells) < 2: continue
            label = row.cells[0].text.strip().lower()
            val_text = row.cells[1].text.strip()
            neg = "(" in val_text and ")" in val_text
            m = re.search(r'\$?([\d,]+\.?\d*)', val_text)
            amt = float(m.group(1).replace(",","")) if m else None
            if amt and neg: amt = -amt
            if "maintenance" in label: data["maintenance"] = amt
            elif "special" in label: data["special_assessments"] = amt
            elif "late" in label: data["late_fees"] = amt
            elif "other charges" in label: data["other_charges"] = amt
            elif "certified" in label: data["certified_mail"] = amt
            elif "other cost" in label: data["other_costs"] = amt
            elif "attorney" in label: data["attorney_fees"] = amt
            elif "partial" in label: data["partial_payment"] = amt
            elif "total" in label:
                data["total"] = amt; total = amt
    return total, data


async def run_test(run_num, entities, matter_label):
    from backend import generate_ground_truth

    print(f"\n{'='*60}")
    print(f"  RUN {run_num} — {matter_label}")
    print(f"{'='*60}")

    # Clear outputs
    for f in Path("outputs").glob("*.xlsx"): f.unlink()
    for f in Path("outputs").glob("*.docx"): f.unlink()

    # Use generate_ground_truth — same endpoint the website calls
    result = await generate_ground_truth(
        certified_mail=40.0, other_costs=16.0, attorney_fees=400.0,
    )
    if result.get("status") != "success":
        raise RuntimeError(f"Generation failed: {result.get('message', result)}")

    ledger_file = Path("outputs") / result["ledger_filename"]
    letter_file = Path("outputs") / result["letter_filename"]
    print(f"  Owner: {result.get('owner')}, Unit: {result.get('unit')}")
    print(f"  Ledger: {result['ledger_filename']}")
    print(f"  Letter: {result['letter_filename']}")

    # Extract & compare
    wb = openpyxl.load_workbook(str(ledger_file))

    soa_total, soa_items = extract_soa(wb)
    nl_total, nl_cats, nola_row, cb_row = extract_nl(wb)

    print(f"\n  SOA TOTAL:       ${soa_total:.2f}" if soa_total else "  SOA TOTAL: NOT FOUND")
    print(f"  NOLA-Ledger NET: ${nl_total:.2f}")
    for k,v in soa_items.items():
        if v and v != 0: print(f"    SOA {k}: ${v:.2f}")
    for k,v in nl_cats.items():
        if v and v != 0: print(f"    NL  {k}: ${v:.2f}")

    checks = []

    # Check 1: SOA vs NOLA-Ledger
    if soa_total:
        d = abs(soa_total - nl_total)
        ok = d < 0.02
        checks.append(("SOA vs NL", ok, soa_total, nl_total, d))
        print(f"\n  CHECK 1 (SOA vs NL): {'PASS' if ok else 'FAIL'} — delta=${d:.2f}")

    # Check 2: Demand letter
    dl_total, dl_data = None, {}
    if letter_file.exists():
        dl_total, dl_data = extract_dl(letter_file)
        print(f"  DL TOTAL:        ${dl_total:.2f}" if dl_total else "  DL TOTAL: NOT FOUND")
        for k,v in dl_data.items():
            if v and v != 0 and k != "total": print(f"    DL  {k}: ${v:.2f}")

        if dl_total and soa_total:
            d2 = abs(soa_total - dl_total)
            ok2 = d2 < 0.02
            checks.append(("SOA vs DL", ok2, soa_total, dl_total, d2))
            print(f"  CHECK 2 (SOA vs DL): {'PASS' if ok2 else 'FAIL'} — delta=${d2:.2f}")

        if dl_total:
            d3 = abs(nl_total - dl_total)
            ok3 = d3 < 0.02
            checks.append(("NL vs DL", ok3, nl_total, dl_total, d3))
            print(f"  CHECK 3 (NL vs DL):  {'PASS' if ok3 else 'FAIL'} — delta=${d3:.2f}")

    # Check 4: DL field mapping — maintenance should not be zero if there are assessments
    if dl_data.get("maintenance") is not None and soa_items.get("nola_opening"):
        nola_open = soa_items["nola_opening"]
        post_asmts = soa_items.get("post_nola_asmts", 0)
        dl_maint = dl_data["maintenance"] or 0
        # DL maintenance should be close to NOLA assessments + post-NOLA assessments
        print(f"  CHECK 4 (field map): DL Maint=${dl_maint:.2f}, SOA NOLA+Asmts=${nola_open + post_asmts:.2f}")

    wb.close()
    passed = all(c[1] for c in checks) if checks else False
    status = "PASS" if passed else ("FAIL" if checks else "ERROR")
    print(f"\n  ── RUN {run_num}: {status} ──")
    return {"status": status, "soa": soa_total, "nl": nl_total, "dl": dl_total,
            "checks": checks, "dl_data": dl_data, "soa_items": soa_items}


async def main():
    # Build entity dicts for each matter from uploaded files
    from backend import _load_entities_from_uploads

    # Get entities (uses files in uploads/)
    entities = await _load_entities_from_uploads()
    owner = entities.get("owner_name", "Unknown")
    unit = entities.get("unit_number", "?")
    assoc = entities.get("association_name", "?")
    print(f"Detected matter: {owner} / Unit {unit} / {assoc}")
    print(f"Key financials from IQ-225:")
    for k in ("principal_balance","late_fees","attorney_fees","other_charges",
              "certified_mail_charges","other_costs","total_amount_owed","monthly_assessment"):
        print(f"  {k}: {entities.get(k, 'N/A')}")

    results = []
    for i in range(1, 6):
        try:
            r = await run_test(i, entities, f"{owner} / Unit {unit}")
            results.append(r)
        except Exception as e:
            print(f"\n  RUN {i} CRASHED: {e}")
            traceback.print_exc()
            results.append({"status": "ERROR", "error": str(e)})

    # Summary
    print(f"\n{'='*60}")
    print(f"  FINAL RESULTS — 5 RUNS")
    print(f"{'='*60}")
    pc = sum(1 for r in results if r["status"]=="PASS")
    fc = sum(1 for r in results if r["status"]=="FAIL")
    ec = sum(1 for r in results if r["status"]=="ERROR")
    for i,r in enumerate(results,1):
        if r["status"]=="ERROR":
            print(f"  Run {i}: ERROR — {r.get('error','?')[:80]}")
        else:
            s=r.get("soa",0)or 0; n=r.get("nl",0)or 0; d=r.get("dl",0)or 0
            print(f"  Run {i}: {r['status']} | SOA=${s:.2f} NL=${n:.2f} DL=${d:.2f}")
            for name,ok,a,b,delta in r.get("checks",[]):
                if not ok: print(f"         FAILED: {name} (${a:.2f} vs ${b:.2f}, delta=${delta:.2f})")
    print(f"\n  PASS: {pc}/5  |  FAIL: {fc}/5  |  ERROR: {ec}/5")
    print(f"{'='*60}")


if __name__ == "__main__":
    asyncio.run(main())
